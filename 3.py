import random
import time
from selenium import webdriver
from selenium.webdriver import ChromeOptions, ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter


chrome_options = webdriver.ChromeOptions()
url = 'https://www.astro.ucla.edu/~wright/CosmoCalc.html'
chrome_options.add_argument('ignore-certificate-errors')
driver = webdriver.Chrome(options=chrome_options)
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
    "source": """
    Object.defineProperty(navigator, 'webdriver', {
      get: () => undefined
    })
  """
})

H = 70
Mm = 0.27
Ma = 0.73


def Grb(z):
    driver.get(url)

    time.sleep(1)
    # driver.find_element_by_id('')
    # driver.find_element_by_xpath('/html/body/form/input[1]').clear()
    # driver.find_element_by_xpath('/html/body/form/input[2]').clear()
    # driver.find_element_by_xpath('/html/body/form/input[3]').clear()
    # driver.find_element_by_xpath('/html/body/form/input[6]').clear()

    # driver.find_element_by_xpath('/html/body/form/input[1]').send_keys(H0)
    # driver.find_element_by_name('txtz').send_keys(H0)
    driver.find_element_by_css_selector("body > form > input[type=TEXT]:nth-child(1)").send_keys(H0)
    driver.find_element_by_xpath('/html/body/form/input[2]').send_keys(z)
    driver.find_element_by_xpath('/html/body/form/input[3]').send_keys(z)
    driver.find_element_by_xpath('/html/body/form/input[5]').click()
    driver.find_element_by_xpath('/html/body/form/input[6]').send_keys(z)
    driver.find_element_by_xpath('/html/body/form/input[7]').click()

    X = driver.find_element_by_xpath('/html/body/ul/li[8]/text()[2]').text

    return X

print(Grb(2))


# 在内存创建一个工作簿obj
# wb = Workbook()
# ws = wb.active
# ws.title = u'Dl'
#
# # 向第一个sheet页写数据吧
# ws.cell(row=1, column=1).value = '名称'
# ws.cell(row=1, column=2).value = 'note0'
#
# for j in range(0, 168):
# # 工作簿保存到磁盘
# wb.save('Dl.xlsx')