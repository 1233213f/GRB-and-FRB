import math
import random
import openpyxl
import numpy as np
from scipy import integrate
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils import get_column_letter, column_index_from_string

# 文档名称
a = '4.xlsx'

# 写入文档的名称
b = '光度.xlsx'

# 定义读取文档为嵌套列表的函数
def dk(a):
    wb = openpyxl.load_workbook(a)  # 打开excel文件
    sheet = wb.active  # 正对表格
    n1 = sheet.max_row
    n2 = sheet.max_column

    data = [[] for _ in range(n1)]
    list = [[] for _ in range(n2)]

    for column in range(1, n1 + 1):
        for row in range(1, n2 + 1):
            list[row - 1] = sheet.cell(column, row).value
        data[column - 1] = list
        list = [[] for _ in range(n2)]
    return data
# print(dk(a), len(dk(a)), len(dk(a)[0]))

def wrt(b):
# 在内存创建一个工作簿
    wb = Workbook()
    ws = wb.active
    for i in range(0, n1):
        ws.cell(row=1, column=11).value = '光度(erg/s)'
        if i < n1 - 1:
            ws.cell(row=i + 2, column=11).value = s[i]
        for j in range(0, n2):
            ws.cell(row=i + 1, column=j + 1).value = d[i][j]

        # 保存文档
    wb.save(b)

# 定义计算光度的函数
def tp(ap, bt, Ep, z, P):

    # 确定分子积分范围
    a = 1 / (1 + z)
    b = 10000 / (1 + z)

    # 确定分母积分范围
    Emin = 15
    Emax = 150

    # 计算分段点
    zz = (ap-bt)*Ep/(2+ap)

    # 定义分段函数
    def ft1(x):
        return x*((x/100)**ap)*np.exp(-(2+ap)*x/Ep)
    def fd1(x):
        return ((x/100)**ap)*np.exp(-(2+ap)*x/Ep)
    def ft2(x):
        return x*(((ap-bt)*Ep/((2+ap)*100))**(ap-bt))*np.exp(bt-ap)*(x/100)**bt
    def fd2(x):
        return (((ap-bt)*Ep/((2+ap)*100))**(ap-bt))*np.exp(bt-ap)*(x/100)**bt

    # 判断分段函数的分段区间，并进行计算
    def top():
        if zz >= Emin and zz <= Emax:
            b1, err11 = integrate.quad(ft1, a, zz)
            c, err12 = integrate.quad(ft2, zz, b)
            e, err21 = integrate.quad(fd1, Emin, zz)
            f, err22 = integrate.quad(fd2, zz, Emax)
            return b1/e + c/f
        elif zz > Emax:
            b1, err11 = integrate.quad(ft1, a, b)
            e, err21 = integrate.quad(fd1, Emin, Emax)
            return b1/e
        elif zz < Emin:
            c, err12 = integrate.quad(ft2, a, b)
            f, err22 = integrate.quad(fd2, Emin, Emax)

    # 1.602 177 33×10**(-16)J， 计算 Pbolo
    Pbolo = 1.60217733 * 10**(-9) * P * top()

    return Pbolo

d = dk(a)
n1 = len(d)
n2 = len(d[0])

# 建立一个元组
s = [[] for _ in range(n1-1)]

# 计算光度
for i in range(0, n1):
    if i + 1 < n1:
        s[i] = 4 * math.pi * tp(d[i + 1][3], d[i + 1][4], d[i + 1][5], d[i + 1][6], d[i + 1][7]) * d[i + 1][8]**2


wrt(b)
# print(s)
# print(wrt(b))
#
# # 在内存创建一个工作簿
# wb = Workbook()
# ws = wb.active
# for i in range(0, n1):
#     ws.cell(row=1, column=11).value = '光度erg/s'
#     if i < n1 - 1:
#         ws.cell(row=i + 1, column=11).value = s[i]
#     for j in range(0, n2):
#         ws.cell(row=i + 1, column=j + 1).value = d[i][j]
#
#     # 保存文档
# wb.save(b)
